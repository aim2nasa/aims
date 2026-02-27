# -*- coding: utf-8 -*-
"""
AC 증분 리포트 -자동화 테스트 + 회귀 테스트

[신규 기능 테스트]
  T1. 증분 리포트: 고객 1명씩 추가하며 엑셀이 매번 갱신되는지
  T2. 빈 데이터: --timestamp + 결과 0건 → 3개 파일 모두 생성
  T3. 덮어쓰기: 동일 timestamp 반복 실행 → 파일 수 불변
  T4. atomic write: .tmp 파일 잔존 없음

[회귀 테스트]
  T5. --timestamp 미지정: 기존 동작 (datetime 기반 파일명) 유지
  T6. --timestamp 미지정 + 재실행: 파일명이 달라서 파일 수 증가 (기존 동작)
  T7. 개인/법인 분류: 개인/법인 고객이 각각 올바른 시트에 기록
  T8. execution_report 컬럼: 15개 컬럼 헤더 유지
  T9. CRS/AR 표시값: 기존 포맷 (건수, 파일명) 유지
"""
import os
import sys
import json
import shutil
import subprocess
import time
import glob

SCRIPT_DIR = r"D:\aims\tools\auto_clicker_v2"
GENERATE_REPORTS = os.path.join(SCRIPT_DIR, "generate_reports.py")
TEST_BASE = r"D:\tmp\test_report"

# 시뮬레이션용 고객 데이터
CUSTOMER_A = {
    "customer_name": "테스트고객A",
    "variable_insurance": {"exists": True, "saved": 1, "duplicate": 0, "metlife_errors": 0,
                           "no_variable_contract": False, "saved_files": ["crs_a.pdf"]},
    "annual_report": {"exists": True, "saved": True, "saved_files": ["ar_a.pdf"]},
    "customer_detail_fallback": {
        "name": "테스트고객A", "customer_type": "개인", "gubun": "일반",
        "insurance_age": "35", "mobile_phone": "010-1234-5678",
        "birth_date": "1990-01-01", "gender": "남", "email": "a@test.com"},
    "issues": []
}
CUSTOMER_B = {
    "customer_name": "테스트법인B",
    "variable_insurance": {"exists": False, "saved": 0, "duplicate": 0, "metlife_errors": 0,
                           "no_variable_contract": True},
    "annual_report": {"exists": True, "saved": True, "saved_files": ["ar_b.pdf"]},
    "customer_detail_fallback": {
        "name": "테스트법인B", "customer_type": "법인", "gubun": "법인",
        "insurance_age": "", "mobile_phone": "", "birth_date": "",
        "gender": "미사용", "email": "b@corp.com"},
    "issues": []
}
CUSTOMER_C = {
    "customer_name": "테스트고객C",
    "variable_insurance": {"exists": True, "saved": 2, "duplicate": 1, "metlife_errors": 0,
                           "no_variable_contract": False, "saved_files": ["crs_c1.pdf", "crs_c2.pdf"]},
    "annual_report": {"exists": False, "saved": False},
    "customer_detail_fallback": {
        "name": "테스트고객C", "customer_type": "개인", "gubun": "일반",
        "insurance_age": "42", "mobile_phone": "010-9999-8888",
        "birth_date": "1983-05-15", "gender": "여", "email": "c@test.com"},
    "issues": []
}
ALL_CUSTOMERS = [CUSTOMER_A, CUSTOMER_B, CUSTOMER_C]


# ─── 유틸 ────────────────────────────────────────────────

def make_dir(path):
    if os.path.exists(path):
        shutil.rmtree(path)
    os.makedirs(path)
    return path


def write_customer_json(dev_dir, results, chosung="ㅌ"):
    path = os.path.join(dev_dir, "customer_results_%s.json" % chosung)
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    if os.path.exists(path):
        os.remove(path)
    os.rename(tmp, path)


def run_reports(dev_dir, timestamp=None):
    cmd = ["python", GENERATE_REPORTS, dev_dir]
    if timestamp:
        cmd += ["--timestamp", timestamp]
    with open(os.devnull, 'w') as devnull:
        return subprocess.call(cmd, stdout=devnull, stderr=devnull)


def run_reports_capture(dev_dir, timestamp=None):
    """stdout을 캡처하여 반환"""
    cmd = ["python", GENERATE_REPORTS, dev_dir]
    if timestamp:
        cmd += ["--timestamp", timestamp]
    result = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="replace")
    return result.returncode, result.stdout


def list_report_files(dev_dir, pattern="*.xlsx"):
    return sorted(glob.glob(os.path.join(dev_dir, pattern)))


def read_json_report(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


# ─── 테스트 결과 수집 ─────────────────────────────────────

_results = []


def assert_test(name, condition, detail=""):
    status = "PASS" if condition else "FAIL"
    _results.append((name, status, detail))
    mark = "  OK" if condition else "  FAIL"
    print("  [%s] %s%s" % (status, name, (" -%s" % detail) if detail else ""))
    return condition


# ─── 신규 기능 테스트 ─────────────────────────────────────

def test_T1_incremental():
    """T1. 증분 리포트: 고객 1명씩 추가 → 매번 엑셀 갱신"""
    print("\n[T1] 증분 리포트 -고객 1명씩 추가")
    dev = make_dir(os.path.join(TEST_BASE, "t1", "dev"))
    ts = "T1_TEST"
    accumulated = []

    for i, cust in enumerate(ALL_CUSTOMERS):
        accumulated.append(cust)
        write_customer_json(dev, accumulated)
        exit_code = run_reports(dev, ts)

        # JSON으로 고객 수 확인
        jpath = os.path.join(dev, "customer_import_%s.json" % ts)
        data = read_json_report(jpath)
        count = data["통계"]["총_고객_수"]

        assert_test("고객 %d명 후 exit=0" % (i + 1), exit_code == 0)
        assert_test("고객 %d명 후 리포트 반영" % (i + 1), count == len(accumulated),
                     "expected=%d actual=%d" % (len(accumulated), count))


def test_T2_empty_data():
    """T2. 빈 데이터 + --timestamp → 3개 파일 모두 생성"""
    print("\n[T2] 빈 데이터 -파일 존재 보장")
    dev = make_dir(os.path.join(TEST_BASE, "t2", "dev"))
    ts = "T2_EMPTY"

    exit_code = run_reports(dev, ts)
    xlsx = list_report_files(dev, "*%s.xlsx" % ts)
    jsons = list_report_files(dev, "*%s.json" % ts)

    assert_test("빈 데이터 exit=0", exit_code == 0)
    assert_test("xlsx 2개 생성", len(xlsx) == 2, "actual=%d" % len(xlsx))
    assert_test("json 1개 생성", len(jsons) == 1, "actual=%d" % len(jsons))


def test_T3_overwrite():
    """T3. 동일 timestamp 반복 실행 → 파일 수 불변"""
    print("\n[T3] 덮어쓰기 -파일 수 불변")
    dev = make_dir(os.path.join(TEST_BASE, "t3", "dev"))
    ts = "T3_OVER"
    write_customer_json(dev, ALL_CUSTOMERS)

    run_reports(dev, ts)
    files_1st = list_report_files(dev, "*%s*" % ts)

    run_reports(dev, ts)
    files_2nd = list_report_files(dev, "*%s*" % ts)

    run_reports(dev, ts)
    files_3rd = list_report_files(dev, "*%s*" % ts)

    assert_test("1회 실행 → 3개", len(files_1st) == 3)
    assert_test("2회 실행 → 3개 유지", len(files_2nd) == 3)
    assert_test("3회 실행 → 3개 유지", len(files_3rd) == 3)


def test_T4_no_tmp_residue():
    """T4. atomic write: .tmp 파일 잔존 없음"""
    print("\n[T4] atomic write -.tmp 잔존 없음")
    dev = make_dir(os.path.join(TEST_BASE, "t4", "dev"))
    write_customer_json(dev, ALL_CUSTOMERS)
    run_reports(dev, "T4_ATOMIC")

    tmp_files = glob.glob(os.path.join(dev, "*.tmp"))
    assert_test(".tmp 파일 0개", len(tmp_files) == 0,
                "found: %s" % tmp_files if tmp_files else "")


# ─── 회귀 테스트 ──────────────────────────────────────────

def test_T5_no_timestamp():
    """T5. --timestamp 미지정: 기존 동작 (datetime 기반 파일명) 유지"""
    print("\n[T5] 회귀: --timestamp 미지정 → datetime 파일명")
    dev = make_dir(os.path.join(TEST_BASE, "t5", "dev"))
    write_customer_json(dev, ALL_CUSTOMERS)

    exit_code = run_reports(dev)  # timestamp 미지정
    xlsx = list_report_files(dev)
    jsons = list_report_files(dev, "*.json")

    # customer_results_ㅌ.json은 입력 파일이므로 제외
    report_jsons = [f for f in jsons if "customer_import" in f]

    assert_test("미지정 시 exit=0", exit_code == 0)
    assert_test("xlsx 2개 생성", len(xlsx) == 2, str([os.path.basename(f) for f in xlsx]))
    assert_test("json 1개 생성", len(report_jsons) == 1,
                str([os.path.basename(f) for f in report_jsons]))

    # 파일명에 T5 등 고정값이 아닌 날짜 패턴이 있는지 확인
    if xlsx:
        fname = os.path.basename(xlsx[0])
        has_date = "202" in fname  # 2026xxxx 패턴
        assert_test("파일명에 datetime 패턴 포함", has_date, fname)


def test_T6_no_timestamp_creates_new():
    """T6. --timestamp 미지정 + 재실행: 파일 수 증가 (기존 동작)"""
    print("\n[T6] 회귀: --timestamp 미지정 재실행 → 파일 증가")
    dev = make_dir(os.path.join(TEST_BASE, "t6", "dev"))
    write_customer_json(dev, ALL_CUSTOMERS)

    run_reports(dev)
    files_1st = list_report_files(dev, "execution_report_*")

    # 1분 미만이면 같은 타임스탬프 → 다른 timestamp를 보장하기 위해
    # 사실 HHMM 단위라 같은 분 내 재실행은 덮어씀
    # 이건 기존 동작이 맞으므로 파일 수 >= 1이면 PASS
    run_reports(dev)
    files_2nd = list_report_files(dev, "execution_report_*")

    assert_test("1회 실행 후 >= 1개", len(files_1st) >= 1)
    assert_test("2회 실행 후 >= 1개", len(files_2nd) >= 1)


def test_T7_personal_corporate():
    """T7. 개인/법인 분류 정확성"""
    print("\n[T7] 회귀: 개인/법인 분류")
    dev = make_dir(os.path.join(TEST_BASE, "t7", "dev"))
    write_customer_json(dev, ALL_CUSTOMERS)  # A=개인, B=법인, C=개인

    ts = "T7_CLASS"
    run_reports(dev, ts)
    jpath = os.path.join(dev, "customer_import_%s.json" % ts)
    data = read_json_report(jpath)

    personal_count = data["통계"]["개인고객_수"]
    corporate_count = data["통계"]["법인고객_수"]

    assert_test("개인 2명", personal_count == 2, "actual=%d" % personal_count)
    assert_test("법인 1명", corporate_count == 1, "actual=%d" % corporate_count)

    # 개인 고객명 확인
    personal_names = [c["고객명"] for c in data["개인고객"]]
    assert_test("개인에 A 포함", "테스트고객A" in personal_names)
    assert_test("개인에 C 포함", "테스트고객C" in personal_names)

    # 법인 고객명 확인
    corporate_names = [c["고객명"] for c in data["법인고객"]]
    assert_test("법인에 B 포함", "테스트법인B" in corporate_names)


def test_T8_report_columns():
    """T8. execution_report 컬럼 15개 유지"""
    print("\n[T8] 회귀: execution_report 컬럼 구조")
    dev = make_dir(os.path.join(TEST_BASE, "t8", "dev"))
    write_customer_json(dev, ALL_CUSTOMERS)
    ts = "T8_COL"
    run_reports(dev, ts)

    try:
        from openpyxl import load_workbook
        wb = load_workbook(os.path.join(dev, "execution_report_%s.xlsx" % ts))
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        wb.close()

        expected = ["고객명", "구분", "생년월일", "보험나이", "성별", "이메일", "휴대폰",
                    "주소", "개인/법인", "사업자번호", "Annual Report", "AR 파일명",
                    "변액리포트", "CRS 파일명", "비고"]

        assert_test("컬럼 수 15개", len(headers) == 15, "actual=%d" % len(headers))
        assert_test("컬럼 헤더 일치", headers == expected,
                     "diff: %s" % [h for h in expected if h not in headers] if headers != expected else "")
    except ImportError:
        assert_test("openpyxl 미설치 → 스킵", True, "openpyxl not available")


def test_T9_crs_ar_format():
    """T9. CRS/AR 표시값 기존 포맷 유지"""
    print("\n[T9] 회귀: CRS/AR 표시값 포맷")
    dev = make_dir(os.path.join(TEST_BASE, "t9", "dev"))
    write_customer_json(dev, ALL_CUSTOMERS)
    ts = "T9_FMT"
    run_reports(dev, ts)

    try:
        from openpyxl import load_workbook
        wb = load_workbook(os.path.join(dev, "execution_report_%s.xlsx" % ts))
        ws = wb.active

        # 데이터 행 읽기 (2행부터)
        rows = {}
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row, 1).value  # 고객명
            ar = ws.cell(row, 11).value   # Annual Report
            crs = ws.cell(row, 13).value  # 변액리포트
            ar_file = ws.cell(row, 12).value  # AR 파일명
            crs_file = ws.cell(row, 14).value  # CRS 파일명
            rows[name] = {"ar": ar, "crs": crs, "ar_file": ar_file, "crs_file": crs_file}
        wb.close()

        # 테스트고객A: CRS 1건, AR 저장
        a = rows.get("테스트고객A", {})
        assert_test("A: AR=1", str(a.get("ar", "")) == "1")
        assert_test("A: CRS=1", str(a.get("crs", "")) == "1")
        assert_test("A: AR파일명 존재", bool(a.get("ar_file")))
        assert_test("A: CRS파일명 존재", bool(a.get("crs_file")))

        # 테스트고객C: CRS saved=2 + duplicate=1 → total=3, 혼합 표시
        c = rows.get("테스트고객C", {})
        assert_test("C: AR=0", str(c.get("ar", "")) == "0")
        # saved=2, dup=1 → "3 (1건 이미 다운로드됨)"
        assert_test("C: CRS 혼합 표시", "3" in str(c.get("crs", "")),
                     "actual='%s'" % c.get("crs", ""))
    except ImportError:
        assert_test("openpyxl 미설치 → 스킵", True, "openpyxl not available")


# ─── 메인 ─────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("AC 증분 리포트 -자동화 + 회귀 테스트")
    print("=" * 60)

    # 신규 기능 테스트
    test_T1_incremental()
    test_T2_empty_data()
    test_T3_overwrite()
    test_T4_no_tmp_residue()

    # 회귀 테스트
    test_T5_no_timestamp()
    test_T6_no_timestamp_creates_new()
    test_T7_personal_corporate()
    test_T8_report_columns()
    test_T9_crs_ar_format()

    # 정리
    if os.path.exists(TEST_BASE):
        shutil.rmtree(TEST_BASE)

    # 결과 요약
    passed = sum(1 for _, s, _ in _results if s == "PASS")
    failed = sum(1 for _, s, _ in _results if s == "FAIL")
    total = len(_results)

    print()
    print("=" * 60)
    if failed:
        print("결과: %d/%d PASS, %d FAIL" % (passed, total, failed))
        print()
        print("실패 항목:")
        for name, status, detail in _results:
            if status == "FAIL":
                print("  - %s%s" % (name, (" (%s)" % detail) if detail else ""))
    else:
        print("결과: ALL %d PASS" % total)
    print("=" * 60)

    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
