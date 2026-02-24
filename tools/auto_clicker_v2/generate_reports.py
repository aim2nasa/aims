# -*- coding: utf-8 -*-
"""
AC v2 리포트 생성기

초성별 customer_results_*.json 파일을 읽어:
1. AIMS 고객일괄등록 호환 엑셀 (customer_import_{timestamp}.xlsx)
2. 동일 내용 JSON (customer_import_{timestamp}.json)
3. 고객별 실행결과 엑셀 (execution_report_{timestamp}.xlsx)

Usage:
    python generate_reports.py <output_base_dir> [--chosung ㄱ,ㄴ,ㄷ]

Examples:
    python generate_reports.py D:\\aims\\tools\\auto_clicker_v2\\output\\ㄴ
    python generate_reports.py D:\\aims\\tools\\auto_clicker_v2\\output\\ㄴ,ㄷ,ㅁ
"""
import os
import re
import sys
import json
import glob
import argparse
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# ──────────────────────────────────────────────────────────
# AIMS 엑셀 포맷 상수 (EXCEL_IMPORT_SPECIFICATION.md v0.2.2)
# ──────────────────────────────────────────────────────────
PERSONAL_COLUMNS = ["고객명", "이메일", "휴대폰", "주소", "성별", "생년월일"]
CORPORATE_COLUMNS = ["고객명", "이메일", "대표전화", "주소", "사업자번호", "대표자명"]

# 실행결과 통합 시트 컬럼
REPORT_COLUMNS = [
    "고객명", "구분", "생년월일", "보험나이", "성별", "이메일", "휴대폰",
    "주소", "개인/법인", "사업자번호",
    "Annual Report", "AR 파일명",
    "변액리포트", "CRS 파일명",
    "비고",
]

# 스타일 상수
HEADER_FONT = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="맑은 고딕", size=10)
DATA_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def load_customer_results(output_dir):
    """초성별 customer_results_*.json 파일 로드 후 통합"""
    all_results = []
    json_files = []

    # 탐색 순서: 직접 → dev/ → 하위 디렉토리 → 하위/dev/
    search_patterns = [
        os.path.join(output_dir, "customer_results_*.json"),
        os.path.join(output_dir, "dev", "customer_results_*.json"),
        os.path.join(output_dir, "*", "customer_results_*.json"),
        os.path.join(output_dir, "*", "dev", "customer_results_*.json"),
    ]
    for pattern in search_patterns:
        json_files = sorted(glob.glob(pattern))
        if json_files:
            break

    for json_path in json_files:
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, list):
                all_results.extend(data)
                print("  [로드] %s → %d명" % (os.path.basename(json_path), len(data)))
        except Exception as e:
            print("  [ERROR] %s 로드 실패: %s" % (json_path, e))

    return all_results


def get_customer_detail(result):
    """customer_result에서 상세정보 추출 (detail + fallback 병합)

    데이터 소스:
      - customer_detail: metdo_reader OCR (고객상세 페이지) → 주소, 사업자번호
      - customer_detail_fallback: 테이블 OCR (고객목록 페이지) → 이메일, 휴대폰, 구분, 보험나이

    병합 규칙: fallback을 기본으로, detail의 비어있지 않은 값으로 덮어쓰기
    → 이메일은 테이블 OCR에서, 주소는 metdo_reader에서 가져옴
    """
    detail = result.get("customer_detail") or {}
    fallback = result.get("customer_detail_fallback") or {}
    # fallback 기본 + detail에서 비어있지 않은 값으로 덮어쓰기
    merged = {}
    merged.update(fallback)
    for k, v in detail.items():
        if v:
            merged[k] = v
    return merged


def normalize_gender(gender_raw):
    """성별 정규화: '남자' → '남', '여자' → '여', '미사용' → ''"""
    if not gender_raw:
        return ""
    g = gender_raw.strip()
    if g in ("남자", "남"):
        return "남"
    if g in ("여자", "여"):
        return "여"
    return ""


def normalize_birth_date(date_raw):
    """생년월일 정규화: 'YYYY.MM.DD' → 'YYYY-MM-DD'"""
    if not date_raw:
        return ""
    return date_raw.strip().replace(".", "-")


def classify_customers(results):
    """고객을 개인/법인으로 분류"""
    personal = []
    corporate = []

    for result in results:
        detail = get_customer_detail(result)
        ctype = detail.get("customer_type", "개인")
        name = result.get("customer_name", "") or detail.get("name", "")

        if not name:
            continue

        if ctype == "법인":
            corporate.append({
                "고객명": name,
                "이메일": detail.get("email", "") or "",
                "대표전화": detail.get("work_phone", "") or detail.get("mobile_phone", "") or "",
                # 법인 주소: 본점소재지(1순위) → 사업장소재지(2순위)
                "주소": detail.get("hq_address", "") or detail.get("business_address", "") or "",
                "사업자번호": detail.get("business_number", "") or "",
                "대표자명": "",  # metdo_reader에서 추출 불가
            })
        else:
            personal.append({
                "고객명": name,
                "이메일": detail.get("email", "") or "",
                "휴대폰": detail.get("mobile_phone", "") or "",
                "주소": detail.get("home_address", "") or detail.get("work_address", "") or "",
                "성별": normalize_gender(detail.get("gender", "")),
                "생년월일": normalize_birth_date(detail.get("birth_date", "")),
            })

    return personal, corporate


def apply_header_style(ws, columns):
    """헤더 행에 스타일 적용"""
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def write_data_rows(ws, data_list, columns):
    """데이터 행 작성"""
    for row_idx, data in enumerate(data_list, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER


def auto_column_width(ws, columns, data_list):
    """컬럼 너비 자동 조절"""
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(col_name) * 2  # 한글은 2배
        for data in data_list[:50]:  # 상위 50개만 샘플링
            val = str(data.get(col_name, ""))
            # 한글 문자는 2배 너비
            char_len = sum(2 if ord(c) > 127 else 1 for c in val)
            max_len = max(max_len, char_len)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)


def generate_customer_import_excel(personal, corporate, output_path):
    """AIMS 고객일괄등록 호환 엑셀 생성 (개인/법인 탭 분리)"""
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    if personal:
        ws = wb.create_sheet("개인고객")
        apply_header_style(ws, PERSONAL_COLUMNS)
        write_data_rows(ws, personal, PERSONAL_COLUMNS)
        auto_column_width(ws, PERSONAL_COLUMNS, personal)

    if corporate:
        ws = wb.create_sheet("법인고객")
        apply_header_style(ws, CORPORATE_COLUMNS)
        write_data_rows(ws, corporate, CORPORATE_COLUMNS)
        auto_column_width(ws, CORPORATE_COLUMNS, corporate)

    if not personal and not corporate:
        print("  [WARN] 개인/법인 고객 데이터가 없습니다.")
        return False

    wb.save(output_path)
    print("  [생성] %s (개인: %d명, 법인: %d명)" % (
        os.path.basename(output_path), len(personal), len(corporate)))
    return True


def generate_customer_import_json(personal, corporate, output_path):
    """AIMS 고객일괄등록 호환 JSON 생성 (엑셀과 동일 내용)"""
    data = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "개인고객": personal,
        "법인고객": corporate,
        "통계": {
            "개인고객_수": len(personal),
            "법인고객_수": len(corporate),
            "총_고객_수": len(personal) + len(corporate),
        }
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print("  [생성] %s" % os.path.basename(output_path))
    return True


# ──────────────────────────────────────────────────────────
# 실행결과 엑셀 (통합 단일 시트)
# ──────────────────────────────────────────────────────────

def format_ar_display(ar_info):
    """Annual Report 표시값과 파일명 반환

    Returns:
        (display_text, filenames_text)
        - display_text: "0", "1", "이미 다운로드됨"
        - filenames_text: 파일명 (쉼표 구분)
    """
    if not ar_info or not ar_info.get("exists"):
        return "0", ""
    if ar_info.get("saved"):
        saved_files = ar_info.get("saved_files", [])
        if saved_files:
            return "1", ", ".join(saved_files)
        # saved=True이지만 새 파일 없음 → 이미 다운로드되어 있었음
        return "이미 다운로드됨", ""
    return "0", ""


def format_crs_display(var_info):
    """변액리포트(CRS) 표시값과 파일명 반환

    Returns:
        (display_text, filenames_text)
        - display_text: "0", "3", "이미 다운로드됨 (2건)", "3 (2건 이미 다운로드됨)"
        - filenames_text: 새로 저장된 파일명 (쉼표 구분)
    """
    if not var_info or not var_info.get("exists"):
        return "0", ""
    saved = var_info.get("saved", 0)
    dup = var_info.get("duplicate", 0)
    total = saved + dup
    saved_files = var_info.get("saved_files", [])
    filenames = ", ".join(saved_files) if saved_files else ""

    if total == 0:
        return "0", ""
    if dup == 0:
        # 모두 새로 저장
        return str(total), filenames
    if saved == 0:
        # 모두 이미 다운로드됨
        if total == 1:
            return "이미 다운로드됨", ""
        return "이미 다운로드됨 (%d건)" % total, ""
    # 혼합: 일부 새로 저장 + 일부 이미 다운로드됨
    return "%d (%d건 이미 다운로드됨)" % (total, dup), filenames


def generate_remarks(result, name_counts, name_seq_num=0):
    """고객 결과에 대한 비고(remarks) 생성

    검사 항목:
      1. 동명이인/중복 계약 (같은 이름 여러 행 → 순번 표시)
      2. 데이터 소스 차이 (detail vs fallback에서 값이 온 경우)
      3. 주소 품질 경고 (파이프, 테이블 데이터 오염)
      4. AR/CRS 특이사항 — 비고 칼럼 단독으로도 주의사항 파악 가능하도록
         AR/CRS 칼럼의 "이미 다운로드됨" 정보를 의도적으로 중복 표시
      5. customer_detail OCR 실패

    Args:
        result: 고객 한 명의 결과 dict
        name_counts: {고객명: 출현횟수} dict (동명이인 감지용)
        name_seq_num: 같은 이름의 몇 번째인지 (1-based, 호출부에서 계산)

    Returns:
        str: 비고 텍스트 (개행 구분)
    """
    notes = []
    detail_raw = result.get("customer_detail") or {}
    fallback_raw = result.get("customer_detail_fallback") or {}
    name = result.get("customer_name", "") or ""

    # 1. 동명이인/중복 계약 (순번 표시)
    if name and name_counts.get(name, 0) > 1:
        total = name_counts[name]
        notes.append("동일인 %d/%d" % (name_seq_num, total))

    # 2. customer_detail OCR 실패
    if not detail_raw:
        notes.append("상세페이지 OCR 실패")

    # 3. 데이터 소스 차이 — 의미 있는 필드만 비교
    source_diffs = []
    # 휴대폰: fallback(테이블)에 없고 detail(상세페이지)에만 있는 경우
    fb_phone = fallback_raw.get("mobile_phone", "") or ""
    dt_phone = detail_raw.get("mobile_phone", "") or ""
    if dt_phone and not fb_phone:
        source_diffs.append("휴대폰: 상세페이지에서 추출")
    elif fb_phone and dt_phone and fb_phone != dt_phone:
        source_diffs.append("휴대폰: 목록=%s, 상세=%s" % (fb_phone, dt_phone))

    # 이메일: detail에만 있고 fallback에 없는 경우 (보통은 fallback에서 옴)
    fb_email = fallback_raw.get("email", "") or ""
    dt_email = detail_raw.get("email", "") or ""
    if dt_email and not fb_email:
        source_diffs.append("이메일: 상세페이지에서 추출")

    # 성별: fallback이 "미사용"인데 detail이 있는 경우
    fb_gender = fallback_raw.get("gender", "") or ""
    dt_gender = detail_raw.get("gender", "") or ""
    if fb_gender == "미사용" and dt_gender and dt_gender != "미사용":
        source_diffs.append("성별: 상세페이지에서 추출 (%s)" % dt_gender)

    if source_diffs:
        notes.extend(source_diffs)

    # 4. 주소 품질 경고 + 주소 소스 추적
    merged = get_customer_detail(result)
    ctype = merged.get("customer_type", "개인")
    if ctype == "법인":
        # 법인: 본점소재지(1순위) → 사업장소재지(2순위)
        address = merged.get("hq_address", "") or merged.get("business_address", "")
        if not merged.get("hq_address") and merged.get("business_address"):
            notes.append("주소: 사업장소재지 사용 (본점소재지 없음)")
    else:
        # 개인: 자택주소(1순위) → 직장주소(2순위)
        address = merged.get("home_address", "") or merged.get("work_address", "")
        if not merged.get("home_address") and merged.get("work_address"):
            notes.append("주소: 직장주소 사용 (자택주소 없음)")
    if address:
        if "|" in address:
            notes.append("주소: 파이프(|) 문자 포함 (OCR 반복 인식 가능성)")
        if len(address) > 150:
            notes.append("주소: 비정상 길이 (%d자)" % len(address))
        # 테이블 데이터 오염 감지 (고객 목록의 이름/전화번호가 주소에 포함)
        if re.search(r'\d{3}-\d{3,4}-\d{4}', address):
            notes.append("주소: 전화번호 패턴 감지 (테이블 데이터 오염 가능)")

    # 5. AR 특이사항
    ar_info = result.get("annual_report", {})
    if ar_info:
        ar_display, _ = format_ar_display(ar_info)
        if ar_display == "이미 다운로드됨":
            notes.append("AR: 이미 다운로드됨")

    # 6. CRS 특이사항
    var_info = result.get("variable_insurance", {})
    if var_info:
        crs_display, _ = format_crs_display(var_info)
        if "이미 다운로드됨" in crs_display:
            notes.append("CRS: %s" % crs_display)

    return "\n".join(notes)


def generate_execution_report(results, output_path):
    """고객별 실행결과 엑셀 생성 (통합 단일 시트)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "실행결과"

    apply_header_style(ws, REPORT_COLUMNS)

    # 동명이인 감지용 이름 카운트
    name_counts = {}
    for result in results:
        n = result.get("customer_name", "") or ""
        if n:
            name_counts[n] = name_counts.get(n, 0) + 1

    # 동명이인 순번 추적 (이름별 등장 순서)
    name_seq = {}

    all_row_data = []
    for row_idx, result in enumerate(results, 2):
        detail = get_customer_detail(result)
        name = result.get("customer_name", "") or detail.get("name", "")
        ctype = detail.get("customer_type", "개인")

        # 동명이인 순번 계산
        if name:
            name_seq[name] = name_seq.get(name, 0) + 1
        seq_num = name_seq.get(name, 1)

        # 주소: 고객 유형별 우선순위
        if ctype == "법인":
            # 법인: 본점소재지(1순위) → 사업장소재지(2순위)
            address = detail.get("hq_address", "") or detail.get("business_address", "")
        else:
            # 개인: 자택주소(1순위) → 직장주소(2순위)
            address = detail.get("home_address", "") or detail.get("work_address", "")

        # 사업자번호: 법인만 해당, 개인은 N/A
        if ctype == "법인":
            biz_num = detail.get("business_number", "") or "-"
        else:
            biz_num = "N/A"

        # AR / CRS
        ar_info = result.get("annual_report", {})
        var_info = result.get("variable_insurance", {})
        ar_display, ar_files = format_ar_display(ar_info)
        crs_display, crs_files = format_crs_display(var_info)

        # 비고 생성
        remarks = generate_remarks(result, name_counts, seq_num)

        row_data = {
            "고객명": name,
            "구분": detail.get("gubun", "") or "-",
            "생년월일": normalize_birth_date(detail.get("birth_date", "")) or "-",
            "보험나이": detail.get("insurance_age", "") or "-",
            "성별": normalize_gender(detail.get("gender", "")) or "-",
            "이메일": detail.get("email", "") or "-",
            "휴대폰": detail.get("mobile_phone", "") or "-",
            "주소": address or "-",
            "개인/법인": ctype,
            "사업자번호": biz_num,
            "Annual Report": ar_display,
            "AR 파일명": ar_files,
            "변액리포트": crs_display,
            "CRS 파일명": crs_files,
            "비고": remarks,
        }
        all_row_data.append(row_data)

        for col_idx, col_name in enumerate(REPORT_COLUMNS, 1):
            value = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER

    auto_column_width(ws, REPORT_COLUMNS, all_row_data)

    # 파일명 컬럼 너비 수동 보정
    ws.column_dimensions["L"].width = 40  # AR 파일명 (12번째)
    ws.column_dimensions["N"].width = 40  # CRS 파일명 (14번째)
    ws.column_dimensions["O"].width = 50  # 비고 (15번째)

    wb.save(output_path)
    print("  [생성] %s (%d명)" % (os.path.basename(output_path), len(results)))
    return True


def main():
    parser = argparse.ArgumentParser(description="AC v2 리포트 생성기")
    parser.add_argument("output_dir", help="출력 베이스 디렉토리 (customer_results_*.json이 있는 경로)")
    parser.add_argument("--chosung", help="초성 필터 (쉼표 구분, 예: ㄱ,ㄴ,ㄷ)")
    args = parser.parse_args()

    output_dir = args.output_dir
    if not os.path.isdir(output_dir):
        print("[ERROR] 디렉토리가 존재하지 않습니다: %s" % output_dir)
        sys.exit(1)

    # 파일명 타임스탬프
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")

    print("=" * 60)
    print("AC v2 리포트 생성기")
    print("=" * 60)
    print("  출력 디렉토리: %s" % output_dir)
    print()

    # 1. customer_results JSON 로드
    print("[1단계] 고객 결과 JSON 로드")
    results = load_customer_results(output_dir)
    if not results:
        print("  [WARN] 로드된 고객 결과가 없습니다.")
        print("  확인: customer_results_*.json 파일이 존재하는지 확인하세요.")
        sys.exit(0)
    print("  총 %d명 로드 완료" % len(results))
    print()

    # 2. 개인/법인 분류
    print("[2단계] 개인/법인 고객 분류")
    personal, corporate = classify_customers(results)
    print("  개인: %d명, 법인: %d명" % (len(personal), len(corporate)))
    print()

    # 3. AIMS 고객일괄등록 엑셀 생성 (개인/법인 탭 분리)
    print("[3단계] AIMS 고객일괄등록 엑셀 생성")
    excel_path = os.path.join(output_dir, "customer_import_%s.xlsx" % timestamp)
    generate_customer_import_excel(personal, corporate, excel_path)
    print()

    # 4. 통합 JSON 생성
    print("[4단계] 통합 JSON 생성")
    json_path = os.path.join(output_dir, "customer_import_%s.json" % timestamp)
    generate_customer_import_json(personal, corporate, json_path)
    print()

    # 5. 고객별 실행결과 엑셀 생성 (통합 단일 시트)
    print("[5단계] 고객별 실행결과 엑셀 생성")
    report_path = os.path.join(output_dir, "execution_report_%s.xlsx" % timestamp)
    generate_execution_report(results, report_path)
    print()

    # 완료
    print("=" * 60)
    print("리포트 생성 완료!")
    print("  - %s" % os.path.basename(excel_path))
    print("  - %s" % os.path.basename(json_path))
    print("  - %s" % os.path.basename(report_path))
    print("=" * 60)


if __name__ == "__main__":
    main()
